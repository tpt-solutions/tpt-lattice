"""Generate a real .xlsx fixture for the tpt-lattice-import-xlsx integration tests.

Run with:  py crates/tpt-lattice-import-xlsx/tests/generate_fixture.py
Requires:  openpyxl
"""
import os

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, Alignment

OUT = os.path.join(os.path.dirname(__file__), "fixtures", "sample.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Numbers, text, boolean.
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3.5
ws["B1"] = "hello"
ws["C1"] = True
ws["C2"] = False

# A formula cell: LES cannot represent it, so the importer must flag it as
# UnsupportedFormula rather than importing a stale cached value.
ws["D1"] = "=A1+A2"

# A defined name referencing a cell (best-effort named-range capture).
wb.defined_names.add(DefinedName(name="MyRange", attr_text="Sheet1!$A$1"))

# --- Styling & merged cells (Phase 5 import surface) -----------------------
# Bold header.
ws["A5"] = "Header"
ws["A5"].font = Font(bold=True)
# Italic note.
ws["A6"] = "Note"
ws["A6"].font = Font(italic=True)
# A number with a custom number format.
ws["A7"] = 3.14159
ws["A7"].number_format = "0.00"
# Horizontally centered text.
ws["A8"] = "centered"
ws["A8"].alignment = Alignment(horizontal="center")

# A merged region; only the top-left cell normally carries a value.
ws["B5"] = "Merged"
ws.merge_cells("B5:C6")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"wrote {OUT}")
